from openpyxl import load_workbook

load_wb = load_workbook("C:/Users/sangm/Downloads/이상민디코100.xlsx", data_only=True)
load_ws = load_wb['Sheet1']
for i in range(1,101):
    load_ws.cell(i, 3).value = f""
